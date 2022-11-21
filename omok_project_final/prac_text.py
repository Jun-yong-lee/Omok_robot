import pygame
import pandas as pd
from openpyxl import load_workbook

from omok_project.button import Button

def get_font(size):  # Returns Press-Start-2P in the desired size
    return pygame.font.Font("tway_sky.ttf", size)

def main():
    screen = pygame.display.set_mode((1280, 800))

    font = pygame.font.Font(None, 32)
    clock = pygame.time.Clock()
    input_box = pygame.Rect(100, 100, 140, 32)
    color_inactive = pygame.Color('lightskyblue3')
    color_active = pygame.Color('dodgerblue2')
    color = color_inactive
    rec = pygame.image.load("image/loading.jpg")
    text = ''
    active = False
    done = False

    data = load_workbook('ranking.xlsx')
    ws = data.active
    name_list = []
    for row in ws.rows:
        name_list.append(row[0].value)

    print(name_list)
    print(type(name_list))
    name_index = 4

    while True:
        screen.fill((0, 0, 0))
        screen.blit(rec, (0, 0))
        MENU_POS = pygame.mouse.get_pos()
        # Render the current text.
        txt_surface = font.render(text, True, color)

        # Resize the box if the text is too long.
        width = max(200, txt_surface.get_width()+10)
        input_box.w = width
        # Blit the text.
        screen.blit(txt_surface, (input_box.x+5, input_box.y+5))
        # Blit the input_box rect.
        pygame.draw.rect(screen, color, input_box, 2)

        add_name_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(450, 500),
                                 text_input="WIN", font=get_font(40), base_color="#d7fcd4",
                                 hovering_color="White")

        back_button = Button(image=pygame.image.load("image/Play Rect.png"), pos=(850, 500),
                             text_input="LOSE", font=get_font(40), base_color="#d7fcd4", hovering_color="White")

        for button in [add_name_button, back_button]:
            button.changeColor(MENU_POS)
            button.update(screen)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                pygame.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if add_name_button.checkForInput(MENU_POS):
                    ws["B"+str((name_index + 1))] = (int(ws["B"+str((name_index + 1))].value) + 1)
                    data.save("ranking.xlsx")
                if back_button.checkForInput(MENU_POS):
                    ws["C"+str((name_index + 1))] = (int(ws["C"+str((name_index + 1))].value) + 1)
                    data.save("ranking.xlsx")

        pygame.display.update()


if __name__ == '__main__':
    pygame.init()
    main()

